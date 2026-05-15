from __future__ import annotations

import io
from collections import defaultdict
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.analytics import (
    Filters, fetch_evaluations, get_filter_options,
    prep_rows, compute_kpi, compute_tab1, compute_tab2, compute_tab3,
    EmployeeFilters, fetch_evaluations_employee, compute_employee_report,
)
from app.database import get_db
from app.deps import get_current_user
from app.models import Block as BlockModel, Checklist, Evaluation, User
from app.scoring import calculate_scores

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from fastapi import Query as QueryParam

router = APIRouter(prefix="/export")

# ── Стили (создаём один раз) ──────────────────────────────────────────────────

_H_FONT  = Font(bold=True, color="FFFFFF", size=10)
_H_FILL  = PatternFill("solid", fgColor="2C3E50")
_H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_FILL_GREEN  = PatternFill("solid", fgColor="D5F5E3")
_FILL_YELLOW = PatternFill("solid", fgColor="FEF9E7")
_FILL_RED    = PatternFill("solid", fgColor="FADBD8")
_FILL_GRAY   = PatternFill("solid", fgColor="F2F3F4")

_FILL_YES = PatternFill("solid", fgColor="D5F5E3")
_FILL_NO  = PatternFill("solid", fgColor="FADBD8")
_FILL_NA  = PatternFill("solid", fgColor="F2F3F4")

_CENTER = Alignment(horizontal="center", vertical="center")


def _score_fill(score):
    if score is None:   return _FILL_GRAY
    if score >= 60:     return _FILL_GREEN
    if score >= 40:     return _FILL_YELLOW
    return _FILL_RED


def _write_header(ws, cols: list[str]):
    ws.append(cols)
    for cell in ws[1]:
        cell.font   = _H_FONT
        cell.fill   = _H_FILL
        cell.alignment = _H_ALIGN
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"


def _autowidth(ws, min_w=8, max_w=38):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        w = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[letter].width = min(max(w + 2, min_w), max_w)


def _result_label(stage):
    return {"сделка успешна": "Успешна",
            "не смог продать": "Не продал",
            "в работе": "В работе"}.get(stage or "", stage or "—")


def _val_label(val):
    return {"yes": "Да", "no": "Нет", "na": "Н/П"}.get(val or "", "—")


# ── Общий запрос ─────────────────────────────────────────────────────────────

def _load_evaluations(db: Session, operator="", checklist_id="", department="",
                      date_from="", date_to="", eval_status="published"):
    q = (
        db.query(Evaluation)
        .options(
            joinedload(Evaluation.evaluator),
            joinedload(Evaluation.items),
            joinedload(Evaluation.checklist).joinedload(Checklist.blocks),
        )
        .filter(Evaluation.status == (eval_status or "published"))
    )
    if operator:
        q = q.filter(Evaluation.operator_name.ilike(f"%{operator}%"))
    if checklist_id:
        q = q.filter(Evaluation.checklist_id == int(checklist_id))
    if department:
        q = q.filter(Evaluation.department == department)
    if date_from:
        try:
            q = q.filter(Evaluation.eval_date >= datetime.strptime(date_from, "%Y-%m-%d").date())
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(Evaluation.eval_date <= datetime.strptime(date_to, "%Y-%m-%d").date())
        except ValueError:
            pass
    evs = q.order_by(Evaluation.eval_date.desc().nullslast(), Evaluation.id.desc()).all()

    # Догружаем критерии (нужны для листа 2)
    cl_ids = {ev.checklist_id for ev in evs if ev.checklist_id}
    if cl_ids:
        full_cls = {
            cl.id: cl for cl in
            db.query(Checklist)
              .filter(Checklist.id.in_(cl_ids))
              .options(joinedload(Checklist.blocks).joinedload(BlockModel.criteria))
              .all()
        }
        for ev in evs:
            if ev.checklist_id in full_cls:
                ev.checklist = full_cls[ev.checklist_id]

    return evs


def _ordered_blocks(evs):
    """[(block_id, checklist_name, block_display_name)] в порядке появления."""
    seen = {}
    for ev in evs:
        if not ev.checklist:
            continue
        for b in ev.checklist.blocks:
            if b.id not in seen:
                seen[b.id] = (b.id, ev.checklist.name, b.display_name or b.name)
    return list(seen.values())


def _multi_cl(blocks):
    return len({cl for _, cl, _ in blocks}) > 1


# ── Лист 1: Детальные оценки ─────────────────────────────────────────────────

def _sheet_detail(wb, evs, blocks):
    ws = wb.create_sheet("Детальные оценки")
    multi = _multi_cl(blocks)
    block_hdrs = [f"{cl} / {bn}" if multi else bn for _, cl, bn in blocks]

    _write_header(ws, [
        "#", "Дата звонка", "Дата оценки", "Сотрудник", "Отдел",
        "Сделка", "Чек-лист", "Оценщик", "Итог %", "Результат", "Комментарий",
    ] + block_hdrs)

    SCORE_COL   = 9
    BLOCK_START = 12

    for ev in evs:
        _, bscores = calculate_scores(ev.items, ev.checklist) if ev.checklist else (None, {})
        block_vals = [bscores.get(bid) for bid, _, _ in blocks]

        ws.append([
            ev.id,
            ev.eval_date.strftime("%d.%m.%Y") if ev.eval_date else "—",
            (ev.updated_at or ev.created_at).strftime("%d.%m.%Y") if (ev.updated_at or ev.created_at) else "—",
            ev.operator_name,
            ev.department or "—",
            f"#{ev.deal_id}" if ev.deal_id else "—",
            ev.checklist.name if ev.checklist else "—",
            (ev.evaluator.full_name or ev.evaluator.username) if ev.evaluator else "—",
            ev.total_score,
            _result_label(ev.stage),
            ev.general_comment or "",
        ] + block_vals)

        ri = ws.max_row
        # Красим только ячейки со скором
        ws.cell(ri, SCORE_COL).fill = _score_fill(ev.total_score)
        ws.cell(ri, SCORE_COL).alignment = _CENTER
        for i, bv in enumerate(block_vals):
            c = ws.cell(ri, BLOCK_START + i)
            c.fill = _score_fill(bv)
            c.alignment = _CENTER

    _autowidth(ws)
    ws.column_dimensions["K"].width = 42


# ── Лист 2: По критериям ─────────────────────────────────────────────────────

def _sheet_criteria(wb, evs):
    ws = wb.create_sheet("По критериям")
    _write_header(ws, [
        "# оценки", "Дата звонка", "Дата оценки", "Сотрудник", "Отдел",
        "Сделка", "Чек-лист", "Оценщик",
        "Блок", "Критерий", "Значение", "Комментарий",
    ])

    VAL_COL = 11

    for ev in evs:
        if not ev.checklist:
            continue
        item_map = {it.criterion_id: it for it in ev.items}
        prefix = [
            ev.id,
            ev.eval_date.strftime("%d.%m.%Y") if ev.eval_date else "—",
            (ev.updated_at or ev.created_at).strftime("%d.%m.%Y") if (ev.updated_at or ev.created_at) else "—",
            ev.operator_name,
            ev.department or "—",
            f"#{ev.deal_id}" if ev.deal_id else "—",
            ev.checklist.name,
            (ev.evaluator.full_name or ev.evaluator.username) if ev.evaluator else "—",
        ]
        for block in ev.checklist.blocks:
            for crit in block.criteria:
                item = item_map.get(crit.id)
                val  = item.value   if item else None
                comm = item.comment if item else ""
                ws.append(prefix + [
                    block.display_name or block.name,
                    crit.text,
                    _val_label(val),
                    comm or "",
                ])
                ri = ws.max_row
                vc = ws.cell(ri, VAL_COL)
                vc.alignment = _CENTER
                if val == "yes":   vc.fill = _FILL_YES
                elif val == "no":  vc.fill = _FILL_NO
                elif val == "na":  vc.fill = _FILL_NA

    _autowidth(ws)
    ws.column_dimensions["J"].width = 48
    ws.column_dimensions["L"].width = 42


# ── Лист 3: Сводка по сотрудникам ────────────────────────────────────────────

def _sheet_summary(wb, evs, blocks):
    ws = wb.create_sheet("Сводка по сотрудникам")
    multi = _multi_cl(blocks)
    block_hdrs = [f"{cl} / {bn}" if multi else bn for _, cl, bn in blocks]

    _write_header(ws, [
        "Сотрудник", "Отдел", "Кол-во оценок", "Средний балл %",
        "Зелёных ≥60%", "Жёлтых 40–60%", "Красных <40%",
    ] + block_hdrs)

    SCORE_COL   = 4
    BLOCK_START = 8

    emp = defaultdict(lambda: {"dept": "—", "scores": [], "bscores": defaultdict(list)})

    for ev in evs:
        if ev.total_score is None or not ev.checklist:
            continue
        k = ev.operator_name
        emp[k]["dept"] = ev.department or "—"
        emp[k]["scores"].append(ev.total_score)
        _, bscores = calculate_scores(ev.items, ev.checklist)
        for bid, s in bscores.items():
            if s is not None:
                emp[k]["bscores"][bid].append(s)

    for name in sorted(emp):
        d = emp[name]
        sc = d["scores"]
        n  = len(sc)
        avg    = round(sum(sc) / n, 1) if n else None
        green  = round(sum(1 for s in sc if s >= 60) / n * 100, 1) if n else 0
        yellow = round(sum(1 for s in sc if 40 <= s < 60) / n * 100, 1) if n else 0
        red    = round(sum(1 for s in sc if s < 40) / n * 100, 1) if n else 0

        bavgs = []
        for bid, _, _ in blocks:
            bs = d["bscores"].get(bid, [])
            bavgs.append(round(sum(bs) / len(bs), 1) if bs else None)

        ws.append([name, d["dept"], n, avg, green, yellow, red] + bavgs)
        ri = ws.max_row
        ws.cell(ri, SCORE_COL).fill = _score_fill(avg)
        ws.cell(ri, SCORE_COL).alignment = _CENTER
        for i, bv in enumerate(bavgs):
            c = ws.cell(ri, BLOCK_START + i)
            c.fill = _score_fill(bv)
            c.alignment = _CENTER

    _autowidth(ws)


# ── Листы для отчётов ────────────────────────────────────────────────────────

def _rpt_sheet_tab1(wb, tab1: dict):
    """Лист 1: Тепловая карта — операторы × блоки."""
    ws = wb.create_sheet("Общие показатели")
    blocks = tab1["blocks"]
    block_names = [b.display_name or b.name for b in blocks]

    _write_header(ws, ["Сотрудник", "Итог %", "% Побед"] + block_names)
    ws.freeze_panes = "A2"

    SCORE_COL   = 2
    BLOCK_START = 4

    for row in tab1["hm_rows"]:
        ws.append([row["name"], row["total"], row["won_pct"]]
                  + [c["pct"] for c in row["cells"]])
        ri = ws.max_row
        ws.cell(ri, SCORE_COL).fill = _score_fill(row["total"])
        ws.cell(ri, SCORE_COL).alignment = _CENTER
        ws.cell(ri, 3).alignment = _CENTER
        for i, c in enumerate(row["cells"]):
            cell = ws.cell(ri, BLOCK_START + i)
            cell.fill = _score_fill(c["pct"])
            cell.alignment = _CENTER

    # Строка команды
    team = tab1["team_cells"]
    ws.append(["Команда (среднее)", tab1["team_total"], tab1.get("team_won_pct")]
              + [c["pct"] for c in team])
    ri = ws.max_row
    for cell in ws[ri]:
        cell.font = Font(bold=True)
    ws.cell(ri, SCORE_COL).fill = _score_fill(tab1["team_total"])
    ws.cell(ri, SCORE_COL).alignment = _CENTER
    for i, c in enumerate(team):
        cell = ws.cell(ri, BLOCK_START + i)
        cell.fill = _score_fill(c["pct"])
        cell.alignment = _CENTER

    # Пустая строка + динамика по неделям
    ws.append([])
    _write_header(ws, ["Неделя", "Кол-во оценок", "Средний балл %"])
    for w in tab1["weekly"]:
        ws.append([w["week"], w["count"], w["avg"]])
        ws.cell(ws.max_row, 3).fill = _score_fill(w["avg"])
        ws.cell(ws.max_row, 3).alignment = _CENTER

    _autowidth(ws)


def _rpt_sheet_tab2(wb, tab2: list):
    """Лист 2: Корреляция — блоки × результат сделки."""
    ws = wb.create_sheet("Корреляция по блокам")
    _write_header(ws, [
        "Блок", "Вес", "Ср. балл (победа) %", "Ср. балл (не продал) %",
        "Разница (Δ) %", "Win rate если сделано %", "Win rate если не сделано %", "Влияние на WR %",
    ])
    ws.freeze_panes = "A2"

    for row in tab2:
        ws.append([
            row["name"], row["weight"],
            row["avg_won"], row["avg_lost"], row["delta"],
            row["wr_done"], row["wr_not_done"], row["wr_impact"],
        ])
        ri = ws.max_row
        ws.cell(ri, 3).fill = _score_fill(row["avg_won"])
        ws.cell(ri, 3).alignment = _CENTER
        ws.cell(ri, 4).fill = _score_fill(row["avg_lost"])
        ws.cell(ri, 4).alignment = _CENTER
        for col in (5, 6, 7, 8):
            ws.cell(ri, col).alignment = _CENTER

    _autowidth(ws)


def _rpt_sheet_tab3(wb, tab3: dict):
    """Лист 3: По сотрудникам — корреляция блоков с результатом."""
    ws = wb.create_sheet("По сотрудникам")
    blocks = tab3["blocks"]
    block_names = [b.display_name or b.name for b in blocks]

    _write_header(ws, ["Сотрудник"] + block_names)
    ws.append(["↓ Победа / Проигрыш →" ] + ["Победа / Проигрыш / Δ"] * len(blocks))
    for cell in ws[ws.max_row]:
        cell.font = Font(italic=True, color="888888", size=8)
    ws.freeze_panes = "A3"

    def _write_op_row(name: str, cells: list):
        won_vals  = [f"{c['won']:.1f}%" if c["won"]  is not None else "—" for c in cells]
        lost_vals = [f"{c['lost']:.1f}%" if c["lost"] is not None else "—" for c in cells]
        delta_vals= [f"{c['delta']:+.1f}%" if c["delta"] is not None else "—" for c in cells]
        combined  = [f"{w} / {l} / {d}" for w, l, d in zip(won_vals, lost_vals, delta_vals)]
        ws.append([name] + combined)
        ri = ws.max_row
        for i, c in enumerate(cells):
            cell = ws.cell(ri, 2 + i)
            cell.alignment = _CENTER
            if c["delta"] is not None:
                cell.fill = _FILL_GREEN if c["delta"] > 5 else (_FILL_RED if c["delta"] < -5 else _FILL_YELLOW)

    for row in tab3["t1_rows"]:
        _write_op_row(row["name"], row["cells"])

    # Команда
    ws.append([])
    _write_op_row("Команда (среднее)", tab3["team_cells"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    _autowidth(ws)
    for i in range(2, len(blocks) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 26


# ── Qolio-совместимый экспорт ─────────────────────────────────────────────────

def _fmt_date_qolio(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d/%m/%Y, %H:%M")


def _qolio_sheet_by_call(wb, evs, selected_cl):
    """По звонкам: строка = оценка, столбцы = метаданные + блоки (0–1) + итог + комм. Формат export(24)."""
    ws = wb.create_sheet("По звонкам")
    blocks = list(selected_cl.blocks)
    block_names = [b.display_name or b.name for b in blocks]

    headers = ["Оператор", "Отдел", "Сделка", "Дата звонка", "Дата оценки"] + \
              block_names + ["Итог", "Комментарий"]
    _write_header(ws, headers)

    BLOCK_START = 6
    TOTAL_COL = BLOCK_START + len(blocks)

    for ev in evs:
        _, bscores = calculate_scores(ev.items, ev.checklist) if ev.checklist else (None, {})
        block_fracs = [round(bscores[b.id] / 100, 3) if bscores.get(b.id) is not None else None for b in blocks]
        total_frac = round(ev.total_score / 100, 3) if ev.total_score is not None else None

        ws.append([
            ev.operator_name,
            ev.department or "",
            ev.deal_id or "",
            _fmt_date_qolio(ev.eval_date),
            _fmt_date_qolio(ev.updated_at or ev.created_at),
        ] + block_fracs + [total_frac, ev.general_comment or ""])

        ri = ws.max_row
        ws.cell(ri, TOTAL_COL).fill = _score_fill(ev.total_score)
        ws.cell(ri, TOTAL_COL).alignment = _CENTER
        for i, b in enumerate(blocks):
            pct = bscores.get(b.id)
            c = ws.cell(ri, BLOCK_START + i)
            c.fill = _score_fill(pct)
            c.alignment = _CENTER

    _autowidth(ws)
    ws.column_dimensions[get_column_letter(len(headers))].width = 42


def _qolio_sheet_by_operator(wb, evs, selected_cl):
    """По сотрудникам: строка = оператор, столбцы = блоки (%), последняя строка = Среднее. Формат export(22)."""
    ws = wb.create_sheet("По сотрудникам")
    blocks = list(selected_cl.blocks)
    block_names = [b.display_name or b.name for b in blocks]

    _write_header(ws, ["Сотрудник", "Отдел", "Пров-ки"] + block_names + ["Итог"])

    BLOCK_START = 4
    TOTAL_COL = BLOCK_START + len(blocks)

    emp: dict = defaultdict(lambda: {"dept": "", "scores": [], "bscores": defaultdict(list)})
    all_scores_g: list[float] = []
    all_bscores_g: dict = defaultdict(list)

    for ev in evs:
        if ev.total_score is None or not ev.checklist:
            continue
        k = ev.operator_name
        emp[k]["dept"] = ev.department or ""
        emp[k]["scores"].append(float(ev.total_score))
        _, bscores = calculate_scores(ev.items, ev.checklist)
        for b in blocks:
            s = bscores.get(b.id)
            if s is not None:
                emp[k]["bscores"][b.id].append(s)
                all_bscores_g[b.id].append(s)
        all_scores_g.append(float(ev.total_score))

    for name in sorted(emp):
        d = emp[name]
        sc = d["scores"]
        n = len(sc)
        avg = round(sum(sc) / n, 1) if n else None
        bavgs = [round(sum(d["bscores"].get(b.id, [])) / len(d["bscores"][b.id]), 1)
                 if d["bscores"].get(b.id) else None for b in blocks]
        ws.append([name, d["dept"], n] + bavgs + [avg])
        ri = ws.max_row
        ws.cell(ri, TOTAL_COL).fill = _score_fill(avg)
        ws.cell(ri, TOTAL_COL).alignment = _CENTER
        for i, bv in enumerate(bavgs):
            c = ws.cell(ri, BLOCK_START + i)
            c.fill = _score_fill(bv)
            c.alignment = _CENTER

    n_all = len(all_scores_g)
    all_avg = round(sum(all_scores_g) / n_all, 1) if n_all else None
    all_bavgs = [round(sum(all_bscores_g[b.id]) / len(all_bscores_g[b.id]), 1)
                 if all_bscores_g.get(b.id) else None for b in blocks]
    ws.append(["Среднее", "", n_all] + all_bavgs + [all_avg])
    ri = ws.max_row
    for cell in ws[ri]:
        cell.font = Font(bold=True)
    ws.cell(ri, TOTAL_COL).fill = _score_fill(all_avg)
    ws.cell(ri, TOTAL_COL).alignment = _CENTER
    for i, bv in enumerate(all_bavgs):
        c = ws.cell(ri, BLOCK_START + i)
        c.fill = _score_fill(bv)
        c.alignment = _CENTER

    _autowidth(ws)


def _qolio_sheet_detailed(wb, evs, selected_cl):
    """По критериям: строка = оценка, пара столбцов на критерий (1/0/пусто + комментарий). Формат export(23)."""
    ws = wb.create_sheet("По критериям")
    blocks = list(selected_cl.blocks)
    all_criteria = [crit for b in blocks for crit in b.criteria]

    meta_h = ["Оператор", "Отдел", "Сделка", "Дата звонка", "Дата оценки"]
    crit_h = [col for crit in all_criteria for col in (crit.text, "Комментарий")]
    tail_h = ["Итог", "Общий комментарий"]

    _write_header(ws, meta_h + crit_h + tail_h)
    ws.row_dimensions[1].height = 60

    CRIT_START = len(meta_h) + 1
    TOTAL_COL = CRIT_START + len(crit_h)

    for ev in evs:
        item_map = {it.criterion_id: it for it in ev.items}
        row: list = [
            ev.operator_name,
            ev.department or "",
            ev.deal_id or "",
            _fmt_date_qolio(ev.eval_date),
            _fmt_date_qolio(ev.updated_at or ev.created_at),
        ]
        for crit in all_criteria:
            item = item_map.get(crit.id)
            if item is None or item.value == "na":
                row.extend([None, (item.comment if item else "") or ""])
            elif item.value == "yes":
                row.extend([1, item.comment or ""])
            else:
                row.extend([0, item.comment or ""])
        total_frac = round(ev.total_score / 100, 3) if ev.total_score is not None else None
        row.extend([total_frac, ev.general_comment or ""])
        ws.append(row)

        ri = ws.max_row
        for i, crit in enumerate(all_criteria):
            item = item_map.get(crit.id)
            c = ws.cell(ri, CRIT_START + i * 2)
            c.alignment = _CENTER
            if item:
                if item.value == "yes":  c.fill = _FILL_YES
                elif item.value == "no": c.fill = _FILL_NO
                else:                    c.fill = _FILL_NA
        ws.cell(ri, TOTAL_COL).fill = _score_fill(ev.total_score)
        ws.cell(ri, TOTAL_COL).alignment = _CENTER

    _autowidth(ws)
    for i in range(len(all_criteria)):
        ws.column_dimensions[get_column_letter(CRIT_START + i * 2 + 1)].width = 30
    ws.column_dimensions[get_column_letter(TOTAL_COL + 1)].width = 42


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("")
def export_xlsx(
    operator: str = "",
    checklist_id: str = "",
    department: str = "",
    date_from: str = "",
    date_to: str = "",
    eval_status: str = "published",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    evs = _load_evaluations(
        db, operator=operator, checklist_id=checklist_id,
        department=department, date_from=date_from, date_to=date_to,
        eval_status=eval_status,
    )
    blocks = _ordered_blocks(evs)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_detail(wb, evs, blocks)
    _sheet_criteria(wb, evs)
    _sheet_summary(wb, evs, blocks)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"callreview_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/reports")
def export_reports_xlsx(
    checklist_id: str = "",
    department: str = "",
    operators: list[str] = QueryParam(default=[]),
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import date as date_type
    from collections import Counter

    filters = Filters(
        department=department,
        operators=operators,
        date_from=date_type.fromisoformat(date_from) if date_from else None,
        date_to=date_type.fromisoformat(date_to) if date_to else None,
        checklist_id=int(checklist_id) if checklist_id else None,
    )

    evaluations = fetch_evaluations(db, filters)
    rows = prep_rows(evaluations)

    # Определяем активный чек-лист (как в reports router)
    cl_counter = Counter(r["ev"].checklist_id for r in rows if r["ev"].checklist_id)
    selected_cl = None
    if filters.checklist_id and cl_counter:
        ids = list(cl_counter.keys())
        cl_map = {
            cl.id: cl for cl in
            db.query(Checklist)
              .options(joinedload(Checklist.blocks))
              .filter(Checklist.id.in_(ids))
              .all()
        }
        selected_cl = cl_map.get(filters.checklist_id)
    if not selected_cl and cl_counter:
        top_id = cl_counter.most_common(1)[0][0]
        selected_cl = (
            db.query(Checklist)
              .options(joinedload(Checklist.blocks))
              .filter(Checklist.id == top_id)
              .first()
        )

    if not selected_cl or not rows:
        # Нет данных — возвращаем пустой файл с пояснением
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Нет данных"
        ws["A1"] = "Нет данных по выбранным фильтрам"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="callreview_reports_empty.xlsx"'},
        )

    rows_for_cl = [r for r in rows if r["ev"].checklist_id == selected_cl.id]

    tab1 = compute_tab1(rows_for_cl, selected_cl)
    tab2 = compute_tab2(rows_for_cl, selected_cl)
    tab3 = compute_tab3(rows_for_cl, selected_cl)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _rpt_sheet_tab1(wb, tab1)
    _rpt_sheet_tab2(wb, tab2)
    _rpt_sheet_tab3(wb, tab3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"callreview_reports_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/qolio")
def export_qolio_xlsx(
    checklist_id: str = "",
    department: str = "",
    operators: list[str] = QueryParam(default=[]),
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import date as date_type
    from collections import Counter

    filters = Filters(
        department=department,
        operators=operators,
        date_from=date_type.fromisoformat(date_from) if date_from else None,
        date_to=date_type.fromisoformat(date_to) if date_to else None,
        checklist_id=int(checklist_id) if checklist_id else None,
    )

    evaluations = fetch_evaluations(db, filters)

    def _empty_response(msg: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Нет данных"
        ws["A1"] = msg
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="qolio_export_empty.xlsx"'},
        )

    if not evaluations:
        return _empty_response("Нет данных по выбранным фильтрам")

    cl_counter = Counter(ev.checklist_id for ev in evaluations if ev.checklist_id)
    selected_cl = None
    if filters.checklist_id:
        selected_cl = next(
            (ev.checklist for ev in evaluations
             if ev.checklist_id == filters.checklist_id and ev.checklist), None
        )
    if not selected_cl and cl_counter:
        top_id = cl_counter.most_common(1)[0][0]
        selected_cl = next(
            (ev.checklist for ev in evaluations
             if ev.checklist_id == top_id and ev.checklist), None
        )

    if not selected_cl:
        return _empty_response("Не удалось определить чек-лист")

    evs_for_cl = [ev for ev in evaluations if ev.checklist_id == selected_cl.id]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _qolio_sheet_by_call(wb, evs_for_cl, selected_cl)
    _qolio_sheet_by_operator(wb, evs_for_cl, selected_cl)
    _qolio_sheet_detailed(wb, evs_for_cl, selected_cl)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"qolio_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Экспорт отчёта «Результаты сотрудников» ──────────────────────────────────

def _emp_sheet_summary(wb, report: dict, display_mode: str):
    ws = wb.create_sheet("Сводная по сотрудникам")
    cols = report["columns"]
    col_labels = [c["label"] for c in cols]
    _write_header(ws, ["Сотрудник"] + col_labels + ["Итог %", "N", "% успех"])

    DATA_START = 2
    TOTAL_COL = len(cols) + 2

    use_pts = display_mode == "pts"

    for row in report["summary_rows"]:
        vals = [
            (c["pts"] if use_pts else c["pct"])
            for c in row["cells"]
        ]
        ws.append([row["name"]] + vals + [row["total"], row["count"], row["won_pct"]])
        ri = ws.max_row
        ws.cell(ri, TOTAL_COL).fill = _score_fill(row["total"])
        ws.cell(ri, TOTAL_COL).alignment = _CENTER
        for i, c in enumerate(row["cells"]):
            cell = ws.cell(ri, DATA_START + i)
            cell.fill = _score_fill(c["pct"])
            cell.alignment = _CENTER

    # Строка команды
    team_vals = [
        (c["pts"] if use_pts else c["pct"])
        for c in report["team_cells"]
    ]
    ws.append(["Команда"] + team_vals + [report["team_total"], report["team_count"], report["team_won_pct"]])
    ri = ws.max_row
    for cell in ws[ri]:
        cell.font = Font(bold=True)
    ws.cell(ri, TOTAL_COL).fill = _score_fill(report["team_total"])
    ws.cell(ri, TOTAL_COL).alignment = _CENTER
    for i, c in enumerate(report["team_cells"]):
        cell = ws.cell(ri, DATA_START + i)
        cell.fill = _score_fill(c["pct"])
        cell.alignment = _CENTER

    _autowidth(ws)


def _emp_sheet_detail(wb, report: dict, display_mode: str):
    ws = wb.create_sheet("По оценкам")
    cols = report["columns"]
    group_mode = report["group_mode"]
    use_pts = display_mode == "pts"

    if group_mode == "criteria":
        col_labels = [f"{c['label']}" for c in cols]
    else:
        col_labels = [c["label"] for c in cols]

    headers = [
        "Сотрудник", "Отдел", "Дата звонка", "Дата оценки",
        "Оценивал",
    ] + col_labels + ["Итог %", "Сделка", "Стадия", "Комментарий"]
    _write_header(ws, headers)
    ws.row_dimensions[1].height = 40

    TOTAL_COL = len(col_labels) + 6
    COL_START = 6

    for row in report["detail_rows"]:
        ev = row["ev"]
        evaluator_name = ""
        if ev.evaluator:
            evaluator_name = ev.evaluator.full_name or ev.evaluator.username

        if group_mode == "criteria":
            cell_vals = []
            for cell in row["cells"]:
                v = cell["value"]
                cell_vals.append(
                    "Да" if v == "yes" else ("Нет" if v == "no" else ("Н/П" if v == "na" else "—"))
                )
        else:
            cell_vals = [
                (c["pts"] if use_pts else c["pct"])
                for c in row["cells"]
            ]

        ws.append([
            ev.operator_name,
            ev.department or "—",
            ev.eval_date.strftime("%d.%m.%Y") if ev.eval_date else "—",
            (ev.updated_at or ev.created_at).strftime("%d.%m.%Y") if (ev.updated_at or ev.created_at) else "—",
            evaluator_name or "—",
        ] + cell_vals + [
            row["total"],
            f"#{ev.deal_id}" if ev.deal_id else "—",
            _result_label(ev.stage),
            ev.general_comment or "",
        ])

        ri = ws.max_row
        ws.cell(ri, TOTAL_COL).fill = _score_fill(row["total"])
        ws.cell(ri, TOTAL_COL).alignment = _CENTER

        for i, (cell_data, raw_cell) in enumerate(zip(row["cells"], row["cells"])):
            c = ws.cell(ri, COL_START + i)
            c.alignment = _CENTER
            if group_mode == "criteria":
                v = cell_data["value"]
                if v == "yes":   c.fill = _FILL_YES
                elif v == "no":  c.fill = _FILL_NO
                elif v == "na":  c.fill = _FILL_NA
            else:
                c.fill = _score_fill(cell_data["pct"])

    _autowidth(ws)
    ws.column_dimensions[get_column_letter(len(headers))].width = 42


@router.get("/employee")
def export_employee_xlsx(
    checklist_id: str = "",
    departments: list[str] = QueryParam(default=[]),
    operators: list[str] = QueryParam(default=[]),
    call_date_from: str = "",
    call_date_to: str = "",
    rated_date_from: str = "",
    rated_date_to: str = "",
    evaluator_id: str = "",
    stage: str = "",
    display_mode: str = "pct",
    group_mode: str = "groups",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import date as date_type

    def _empty(msg: str):
        wb2 = openpyxl.Workbook()
        wb2.active.title = "Нет данных"
        wb2.active["A1"] = msg
        buf2 = io.BytesIO()
        wb2.save(buf2)
        buf2.seek(0)
        return StreamingResponse(
            buf2,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="employee_report_empty.xlsx"'},
        )

    if not checklist_id:
        return _empty("Чек-лист не выбран")

    selected_cl = (
        db.query(Checklist)
        .options(joinedload(Checklist.blocks).joinedload(BlockModel.criteria))
        .filter(Checklist.id == int(checklist_id))
        .first()
    )
    if not selected_cl:
        return _empty("Чек-лист не найден")

    filters = EmployeeFilters(
        departments=departments,
        operators=operators,
        call_date_from=date_type.fromisoformat(call_date_from) if call_date_from else None,
        call_date_to=date_type.fromisoformat(call_date_to) if call_date_to else None,
        rated_date_from=date_type.fromisoformat(rated_date_from) if rated_date_from else None,
        rated_date_to=date_type.fromisoformat(rated_date_to) if rated_date_to else None,
        checklist_id=int(checklist_id),
        evaluator_id=int(evaluator_id) if evaluator_id else None,
        stage=stage or None,
        group_mode=group_mode if group_mode in ("groups", "criteria") else "groups",
        display_mode=display_mode if display_mode in ("pct", "pts") else "pct",
    )

    evaluations = fetch_evaluations_employee(db, filters)
    if not evaluations:
        return _empty("Нет данных по выбранным фильтрам")

    report = compute_employee_report(evaluations, selected_cl, filters.group_mode)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _emp_sheet_summary(wb, report, display_mode)
    _emp_sheet_detail(wb, report, display_mode)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    cl_name = selected_cl.name.replace(" ", "_")[:30]
    fname = f"employees_{cl_name}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    fname_ascii = f"employees_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    encoded = quote(fname, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{fname_ascii}\"; filename*=UTF-8''{encoded}"},
    )
